# -*- coding: utf-8 -*-
# @Time ： 2020/7/15 14:01
# @Auth ： tudou
# @File ：lesson7.py
# @QQ ：121313927
# @Company：湖南省零檬信息技术有限公司

"""
接口自动化步骤：
1、excel测试用例准备ok，代码自动读取测试数据 -- read_data()
2、发送接口请求，得到响应信息   -- api_fun()
3、断言：实际结果 vs 预期结果 -- 通过/不通过   --- 断言
4、写入通过/不通过 -- excel  --write_result()

--eval()--运行被字符串包裹的表达式
'{"mobile_phone":"13652440102","pwd":"lemon12345678901","type":1,"reg_name":"6"}'
’5*6‘   --30
"""

import requests
import openpyxl

# 读取测试用例函数
def read_data(filename,sheetname):
    wb = openpyxl.load_workbook(filename)   # 加载工作簿 -- 文档名字
    sheet = wb[sheetname]   # 获取表单
    max_row = sheet.max_row  # 获取最大行数
    case_list = [] # 创建空列表，存放测试用例
    for i in range(2,max_row+1):
        dict1 = dict(
        case_id = sheet.cell(row=i, column=1).value,  # 获取case_id
        url = sheet.cell(row=i,column=5).value,   # 获取Url
        data = sheet.cell(row=i,column=6).value,  # 获取data
        expect = sheet.cell(row=i,column=7).value, # 获取expect
        )
        case_list.append(dict1)  # 每循环一次，就把读取到的字典数据存放到这个list
    return case_list   # 返回测试用例列表

# 执行接口函数
def api_fun(url,data):
    headers_reg = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}  # 请求头- 字典
    res = requests.post(url=url,json=data,headers=headers_reg)  # 接收post方法的结果
    response = res.json()  # 响应正文
    return response

# 写入结果
def write_result(filename,sheetname,row,column,final_result):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    sheet.cell(row=row, column=column).value = final_result  # 写入结果
    wb.save(filename)  # 保存，关闭文档

# 执行测试用例并回写实际结果
def execute_fun(filename,sheetname):
    cases = read_data(filename,sheetname)   # 调用读取测试用例，获取所有测试用例数据保存到变量
    for case in cases:
        case_id = case.get('case_id')  # case['case_id']
        url = case.get('url')
        data = eval(case.get('data'))  # eval() 运行被字符串包裹的表达式----去掉字符串引号
        expect = eval(case.get('expect'))  # 获取预期结果
        expect_msg = expect.get('msg')   # 获取预期结果中的msg
        real_result = api_fun(url=url,data=data)    # 调用发送接口请求函数,返回结果用变量real_result接收
        real_msg = real_result.get('msg')  # 获取实际结果中的msg
        print('预期结果中的msg：{}'.format(expect_msg))
        print('实际结果中的msg：{}'.format(real_msg))
        if real_msg==expect_msg:
            print('第{}条测试用例执行通过！'.format(case_id))
            final_re = 'Passed'
        else:
            print('第{}条测试用例执行不通过！'.format(case_id))
            final_re = 'Failed'
        write_result(filename,sheetname,case_id+1,8,final_re)
        print("*"*25)

execute_fun('test_case_api.xlsx','login')


